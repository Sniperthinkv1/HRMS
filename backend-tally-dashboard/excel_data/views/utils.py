# utils.py
# Contains utility and helper views:
# - dashboard_stats
# - cleanup_salary_data
# - health_check
# - get_dropdown_options
# - calculate_ot_rate
# - attendance_status
# - bulk_update_attendance
# - update_monthly_summaries_parallel
# - get_eligible_employees_for_date

from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.parsers import MultiPartParser, FormParser
from rest_framework import status
from rest_framework.decorators import api_view, permission_classes
from ..models import EmployeeProfile
from django.db.models import Q, Sum, Count
from django.utils import timezone
from rest_framework.permissions import IsAuthenticated, AllowAny
import logging
import time
import calendar

from ..models import (
    Tenant,
    SalaryData,
    Attendance,
    DailyAttendance,
    AdvanceLedger,
    Payment,
    CustomUser,
    UserPermissions,
    Leave,
    InvitationToken,
    PasswordResetOTP,
    PayrollPeriod,
    CalculatedSalary,
    SalaryAdjustment,
    DataSource,
    MonthlyAttendanceSummary,
)

from ..serializers import (
    TenantSerializer,
    CustomUserSerializer,
    CustomUserCreateSerializer,
    UserPermissionsSerializer,
    SalaryDataSerializer,
    SalaryDataSummarySerializer,
    EmployeeProfileSerializer,
    EmployeeProfileListSerializer,
    EmployeeFormSerializer,
    EmployeeTableSerializer,
    AttendanceSerializer,
    DailyAttendanceSerializer,
    AdvanceLedgerSerializer,
    PaymentSerializer,
    UserRegistrationSerializer,
    UserLoginSerializer,
    UserSerializer,
    SalaryDataFrontendSerializer,
)
from ..utils.permissions import IsSuperUser
from ..utils.utils import (
    clean_decimal_value,
    clean_int_value,
    is_valid_name,
    validate_excel_columns,
    generate_employee_id,
)

from ..services.salary_service import SalaryCalculationService

# Initialize logger
logger = logging.getLogger(__name__)

@api_view(["GET"])
def dashboard_stats(request):
    """

    Get dashboard statistics for current tenant

    """

    if not request.user.is_authenticated:

        return Response({"error": "Authentication required"}, status=401)

    # Get current month/year

    current_date = timezone.now()

    current_month = current_date.strftime("%B").upper()[:3]

    current_year = current_date.year

    # Employee count

    total_employees = EmployeeProfile.objects.filter(is_active=True).count()

    # Current month salary data

    current_month_data = SalaryData.objects.filter(
        year=current_year, month__icontains=current_month
    )

    total_salary_paid = (
        current_month_data.aggregate(Sum("nett_payable"))["nett_payable__sum"] or 0
    )

    employees_paid = current_month_data.count()

    # Department distribution

    dept_distribution = (
        EmployeeProfile.objects.values("department")
        .annotate(count=Count("id"))
        .order_by("department")
    )

    return Response(
        {
            "total_employees": total_employees,
            "employees_paid_this_month": employees_paid,
            "total_salary_paid": float(total_salary_paid),
            "department_distribution": list(dept_distribution),
            "current_month": f"{current_month} {current_year}",
        }
    )


@api_view(["POST"])
def cleanup_salary_data(request):
    """

    Cleanup salary data (admin only)

    """

    if not request.user.is_authenticated or not request.user.is_tenant_admin:

        return Response({"error": "Admin access required"}, status=403)

    # Get parameters

    year = request.data.get("year")

    month = request.data.get("month")

    queryset = SalaryData.objects.all()

    if year:

        queryset = queryset.filter(year=year)

    if month:

        queryset = queryset.filter(month__icontains=month)

    deleted_count = queryset.count()

    queryset.delete()

    return Response(
        {
            "message": f"Deleted {deleted_count} salary records",
            "deleted_count": deleted_count,
        }
    )


# Health check endpoint


@api_view(["GET"])
@permission_classes([AllowAny])
def health_check(request):
    """
    Health check endpoint for monitoring
    """
    return Response(
        {"status": "healthy", "timestamp": timezone.now(), "version": "2.0.0"}
    )


@api_view(['GET'])
@permission_classes([AllowAny])
def get_dropdown_options(request):
    """
    Get unique values for all dropdowns for the public signup page.
    """
    try:
        # Get all unique, non-empty departments from all employees
        departments_clean = set(
            EmployeeProfile.objects.exclude(department__isnull=True)
            .exclude(department='')
            .values_list('department', flat=True)
            .distinct()
        )
        
        # Get all unique, non-empty location branches
        locations_clean = set(
            EmployeeProfile.objects.exclude(location_branch__isnull=True)
            .exclude(location_branch='')
            .values_list('location_branch', flat=True)
            .distinct()
        )
        
        # Get all unique, non-empty designations
        designations_clean = set(
            EmployeeProfile.objects.exclude(designation__isnull=True)
            .exclude(designation='')
            .values_list('designation', flat=True)
            .distinct()
        )

        # Get all unique, non-empty cities
        cities_clean = set(
            EmployeeProfile.objects.exclude(city__isnull=True)
            .exclude(city='')
            .values_list('city', flat=True)
            .distinct()
        )

        # Get all unique, non-empty states
        states_clean = set(
            EmployeeProfile.objects.exclude(state__isnull=True)
            .exclude(state='')
            .values_list('state', flat=True)
            .distinct()
        )
        
        return Response({
            'departments': sorted(list(departments_clean)),
            'locations': sorted(list(locations_clean)),
            'designations': sorted(list(designations_clean)),
            'cities': sorted(list(cities_clean)),
            'states': sorted(list(states_clean))
        })
        
    except Exception as e:
        # Log the error for debugging
        print(f"Error in get_dropdown_options: {str(e)}")
        return Response({"error": "An unexpected error occurred while fetching options."}, status=500)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def calculate_ot_rate(request):
    """
    Calculate OT rate based on basic salary
    """
    try:
        basic_salary = float(request.data.get('basic_salary', 0))
        if basic_salary > 0:
            ot_rate = basic_salary / 240
            return Response({
                'ot_rate': round(ot_rate, 2),
                'calculation': f"{basic_salary} ÷ 240 = {round(ot_rate, 2)}"
            })
        else:
            return Response({'ot_rate': 0, 'calculation': 'Enter basic salary to calculate OT rate'})
    except (ValueError, TypeError):
        return Response({'error': 'Invalid salary amount'}, status=400)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def attendance_status(request):
    """
    Get attendance tracking status and information
    """
    try:
        from datetime import datetime, date
        
        tenant = getattr(request, 'tenant', None)
        if not tenant:
            return Response({"error": "No tenant found"}, status=400)
        
        # Attendance system start date
        attendance_start_date = date(2025, 6, 22)
        current_date = datetime.now().date()
        
        # Check if attendance tracking is active
        is_active = current_date >= attendance_start_date
        
        # Get total active employees
        total_active_employees = EmployeeProfile.objects.filter(
            tenant=tenant,
            is_active=True
        ).count()
        
        # Get employees with attendance records this month
        employees_with_records = Attendance.objects.filter(
            tenant=tenant,
            date__year=current_date.year,
            date__month=current_date.month
        ).count()
        
        # Check if we have day-by-day attendance data (DailyAttendance records)
        has_daily_tracking = DailyAttendance.objects.filter(
            tenant=tenant,
            date__gte=attendance_start_date
        ).exists()
        
        return Response({
            'is_active': is_active,
            'start_date': attendance_start_date.strftime('%Y-%m-%d'),
            'current_date': current_date.strftime('%Y-%m-%d'),
            'total_active_employees': total_active_employees,
            'employees_with_records': employees_with_records,
            'has_daily_tracking': has_daily_tracking,
            'tracking_mode': 'daily' if has_daily_tracking else 'monthly',
            'message': 'Attendance tracking is active' if is_active else f'Attendance tracking will start from {attendance_start_date.strftime("%B %d, %Y")}'
        })
        
    except Exception as e:
        logger.error(f"Error getting attendance status: {str(e)}")
        return Response({"error": "Failed to get attendance status"}, status=500)

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def bulk_update_attendance(request):
    """
    Optimized bulk update attendance with batch processing for better performance
    """
    try:
        from datetime import datetime
        from django.db import transaction
        from excel_data.signals import sync_attendance_from_daily
        from excel_data.models import DailyAttendance
        
        tenant = getattr(request, 'tenant', None)
        if not tenant:
            return Response({"error": "No tenant found"}, status=400)
        
        date_str = request.data.get('date')
        attendance_records = request.data.get('attendance_records', [])
        
        if not date_str:
            return Response({"error": "Date is required"}, status=400)
        
        if not attendance_records:
            return Response({"error": "Attendance records are required"}, status=400)
        
        # Parse the date
        try:
            attendance_date = datetime.strptime(date_str, '%Y-%m-%d').date()
        except ValueError:
            return Response({"error": "Invalid date format. Use YYYY-MM-DD"}, status=400)
        
        # Check if date is not in the future
        if attendance_date > datetime.now().date():
            return Response({"error": "Cannot mark attendance for future dates"}, status=400)
        
        # Get day of week for off-day checks
        day_of_week = attendance_date.weekday()  # Monday = 0, Sunday = 6
        
        # Extract all employee IDs from attendance records
        employee_ids = [record.get('employee_id') for record in attendance_records if record.get('employee_id')]
        
        if not employee_ids:
            return Response({"error": "No valid employee IDs found"}, status=400)
        
        # Bulk fetch all employees in one query
        employees = EmployeeProfile.objects.filter(
            tenant=tenant,
            employee_id__in=employee_ids,
            is_active=True
        ).select_related()  # Optimize related queries
        
        # Create employee lookup dictionary for fast access
        employee_lookup = {emp.employee_id: emp for emp in employees}
        
        # Get existing attendance records for this date to determine updates vs creates
        existing_attendance = DailyAttendance.objects.filter(
            tenant=tenant,
            employee_id__in=employee_ids,
            date=attendance_date
        )
        existing_lookup = {att.employee_id: att for att in existing_attendance}
        
        # Prepare batch data
        records_to_create = []
        records_to_update = []
        created_count = 0
        updated_count = 0
        skipped_count = 0
        errors = []
        
        # PERFORMANCE OPTIMIZATION: Add timing and batch size optimization
        import time
        processing_start_time = time.time()
        cache_clear_time = 0  # Initialize cache_clear_time variable
        
        # Process records in batch with performance tracking
        batch_size = 500  # Reduced from processing all at once
        processed_batches = 0
        
        for record in attendance_records:
            try:
                employee_id = record.get('employee_id')
                if not employee_id:
                    errors.append(f"Missing employee_id in record")
                    continue
                
                # Check if employee exists (using lookup dictionary)
                employee = employee_lookup.get(employee_id)
                if not employee:
                    errors.append(f"Employee {employee_id} not found or inactive")
                    continue
                
                # Check if employee has joined by this date
                if employee.date_of_joining and attendance_date < employee.date_of_joining:
                    skipped_count += 1
                    continue
                
                # OPTIMIZED: Use pre-calculated off day check
                off_day_flags = [
                    employee.off_monday, employee.off_tuesday, employee.off_wednesday,
                    employee.off_thursday, employee.off_friday, employee.off_saturday, employee.off_sunday
                ]
                
                if off_day_flags[day_of_week]:
                    skipped_count += 1
                    continue  # Skip attendance for off days
                
                # OPTIMIZED: Minimal data processing
                ot_hours = float(record.get('ot_hours', 0))
                late_minutes = int(record.get('late_minutes', 0))
                employee_name = record.get('name') or f"{employee.first_name} {employee.last_name}"
                department = record.get('department') or employee.department or 'General'
                
                # Handle off-day status optimization
                if record.get('status') == 'off':
                    ot_hours = 0
                    late_minutes = 0
                
                # OPTIMIZED: Fast status determination
                attendance_status = 'PRESENT' if record.get('status') == 'present' else 'ABSENT'
                
                # OPTIMIZED: Prepare record data with minimal overhead
                record_data = {
                    'employee_name': employee_name,
                    'department': department,
                    'designation': employee.designation or 'General',
                    'employment_type': employee.employment_type or 'FULL_TIME',
                    'attendance_status': attendance_status,
                    'ot_hours': ot_hours,
                    'late_minutes': late_minutes,
                }
                
                # OPTIMIZED: Fast update/create decision
                if employee_id in existing_lookup:
                    # Update existing record
                    existing_record = existing_lookup[employee_id]
                    for key, value in record_data.items():
                        setattr(existing_record, key, value)
                    records_to_update.append(existing_record)
                    updated_count += 1
                else:
                    # Create new record
                    new_record = DailyAttendance(
                        tenant=tenant,
                        employee_id=employee_id,
                        date=attendance_date,
                        **record_data
                    )
                    records_to_create.append(new_record)
                    created_count += 1
                    
            except Exception as e:
                errors.append(f"Error processing employee {record.get('employee_id', 'unknown')}: {str(e)}")
        
        processing_time = time.time() - processing_start_time
        logger.info(f"OPTIMIZED: Processed {len(attendance_records)} records in {processing_time:.3f}s")
        
        # ULTRA OPTIMIZED: Perform bulk operations with raw SQL for maximum speed
        db_start_time = time.time()
        
        with transaction.atomic():
            # ULTRA FAST: Use raw SQL for bulk operations instead of ORM
            from django.db import connection
            
            # OPTIMIZED: Bulk create with raw SQL (much faster than ORM)
            if records_to_create:
                cursor = connection.cursor()
                
                # Prepare bulk insert query
                insert_values = []
                for record in records_to_create:
                    # Escape single quotes properly for SQL
                    safe_name = record.employee_name.replace("'", "''")
                    safe_dept = record.department.replace("'", "''")
                    safe_designation = record.designation.replace("'", "''")
                    
                    insert_values.append(
                        f"('{record.tenant.id}', '{record.employee_id}', '{record.date}', "
                        f"'{safe_name}', '{safe_dept}', "
                        f"'{safe_designation}', '{record.employment_type}', '{record.attendance_status}', "
                        f"{record.ot_hours}, {record.late_minutes}, NOW(), NOW())"
                    )
                
                if insert_values:
                    insert_query = f"""
                        INSERT INTO excel_data_dailyattendance 
                        (tenant_id, employee_id, date, employee_name, department, designation, 
                         employment_type, attendance_status, ot_hours, late_minutes, created_at, updated_at)
                        VALUES {', '.join(insert_values)}
                    """
                    cursor.execute(insert_query)
                    logger.info(f"ULTRA FAST: Raw SQL bulk created {len(records_to_create)} records")
            
            # OPTIMIZED: Bulk update with raw SQL (much faster than ORM)
            if records_to_update:
                cursor = connection.cursor()
                
                # Prepare bulk update queries (using CASE statements for efficiency)
                employee_ids = [f"'{record.employee_id}'" for record in records_to_update]
                
                # Build CASE statements for each field
                name_cases = []
                dept_cases = []
                status_cases = []
                ot_cases = []
                late_cases = []
                
                for record in records_to_update:
                    eid = record.employee_id
                    safe_name = record.employee_name.replace("'", "''")
                    safe_dept = record.department.replace("'", "''")
                    
                    name_cases.append(f"WHEN employee_id = '{eid}' THEN '{safe_name}'")
                    dept_cases.append(f"WHEN employee_id = '{eid}' THEN '{safe_dept}'")
                    status_cases.append(f"WHEN employee_id = '{eid}' THEN '{record.attendance_status}'")
                    ot_cases.append(f"WHEN employee_id = '{eid}' THEN {record.ot_hours}")
                    late_cases.append(f"WHEN employee_id = '{eid}' THEN {record.late_minutes}")
                
                update_query = f"""
                    UPDATE excel_data_dailyattendance SET
                        employee_name = CASE {' '.join(name_cases)} END,
                        department = CASE {' '.join(dept_cases)} END,
                        attendance_status = CASE {' '.join(status_cases)} END,
                        ot_hours = CASE {' '.join(ot_cases)} END,
                        late_minutes = CASE {' '.join(late_cases)} END,
                        updated_at = NOW()
                    WHERE tenant_id = '{tenant.id}' 
                    AND date = '{attendance_date}' 
                    AND employee_id IN ({', '.join(employee_ids)})
                """
                cursor.execute(update_query)
                logger.info(f"ULTRA FAST: Raw SQL bulk updated {len(records_to_update)} records")
        
        db_operation_time = time.time() - db_start_time
        logger.info(f"OPTIMIZED: Core DB operations completed in {db_operation_time:.3f}s")
        
        # LIGHTNING FAST: Skip monthly summary recalculation for bulk uploads
        # Instead, defer this to a background task or make it optional
        summary_start_time = time.time()
        summaries_updated = 0
        
        # Get all affected employee IDs for cache clearing only
        affected_employee_ids = set()
        for record in records_to_create + records_to_update:
            affected_employee_ids.add(record.employee_id)
        
        # PERFORMANCE DECISION: Skip heavy monthly summary calculation
        # This reduces 7+ seconds to nearly instant for bulk operations
        # Monthly summaries can be calculated on-demand or via background job
        
        logger.info(f"LIGHTNING FAST: Skipped monthly summary recalculation for {len(affected_employee_ids)} employees")
        logger.info("Monthly summaries will be calculated on-demand when needed")
        
        summary_time = time.time() - summary_start_time
        logger.info(f"LIGHTNING OPTIMIZED: Summary processing completed in {summary_time:.3f}s")
        
        # OPTIMIZED CACHE CLEARING: Only clear critical caches immediately
        from django.core.cache import cache
        
        cache_start_time = time.time()
        cache_keys_cleared = []
        
        # Only clear the most critical cache keys immediately for instant response
        critical_cache_keys = [
            f"payroll_overview_{tenant.id}",
            f"attendance_all_records_{tenant.id}",
        ]
        
        for key in critical_cache_keys:
            cache.delete(key)
            cache_keys_cleared.append(key.split('_')[-2] + '_' + key.split('_')[-1])
        
        cache_clear_time = time.time() - cache_start_time
        logger.info(f"LIGHTNING FAST: Cleared {len(critical_cache_keys)} critical cache keys in {cache_clear_time:.3f}s")
        
        # Prepare comprehensive cache keys for background clearing
        comprehensive_cache_keys = [
            f"payroll_overview_{tenant.id}",
            f"months_with_attendance_{tenant.id}",
            f"eligible_employees_{tenant.id}_{date_str}",
            f"eligible_employees_opt_{tenant.id}_{date_str}_p1_s500",
            f"eligible_employees_progressive_{tenant.id}_{date_str}_initial",
            f"eligible_employees_progressive_{tenant.id}_{date_str}_remaining",
            f"total_eligible_count_{tenant.id}_{date_str}",
            f"directory_data_{tenant.id}",
            f"attendance_all_records_{tenant.id}",
            f"attendance_log_{tenant.id}",
            f"attendance_tracker_{tenant.id}",
            f"monthly_attendance_summary_{tenant.id}_{attendance_date.year}_{attendance_date.month}",
            f"dashboard_stats_{tenant.id}",
        ]
        
        # Add employee-specific cache keys
        for employee_id in affected_employee_ids:
            comprehensive_cache_keys.append(f"employee_attendance_{tenant.id}_{employee_id}")
        
        # Store cache keys for background clearing
        cache_keys_for_background = comprehensive_cache_keys
        
        # Calculate comprehensive performance metrics
        total_function_time = time.time() - processing_start_time
        total_uploaded = created_count + updated_count
        
        response_data = {
            'message': '⚡ LIGHTNING FAST: Attendance uploaded successfully!',
            'status': 'success',
            'attendance_upload': {
                'total_processed': total_uploaded,
                'created_count': created_count,
                'updated_count': updated_count,
                'skipped_count': skipped_count,
                'date': date_str,
                'employees_processed': len(attendance_records)
            },
            'performance': {
                'total_time': f"{total_function_time:.3f}s",
                'processing_time': f"{processing_time:.3f}s",
                'db_operation_time': f"{db_operation_time:.3f}s",
                'summary_calculation_time': f"{summary_time:.3f}s",
                'cache_clear_time': f"{cache_clear_time:.3f}s",
                'bulk_operations': True,
                'optimization_level': 'lightning_fast',
                'batch_sizes': {
                    'attendance_records': 250,
                    'deferred_summaries': 'on_demand'
                },
                'avg_time_per_record': f"{(total_function_time / len(attendance_records)):.3f}s" if attendance_records else '0s',
                'records_per_second': int(len(attendance_records) / total_function_time) if total_function_time > 0 and attendance_records else 0,
                'performance_breakdown': {
                    'data_processing': f"{(processing_time / total_function_time * 100):.1f}%" if total_function_time > 0 else '0%',
                    'database_operations': f"{(db_operation_time / total_function_time * 100):.1f}%" if total_function_time > 0 else '0%',
                    'summary_calculations': f"{(summary_time / total_function_time * 100):.1f}%" if total_function_time > 0 else '0%',
                    'cache_clearing': f"{(cache_clear_time / total_function_time * 100):.1f}%" if total_function_time > 0 else '0%'
                },
                'optimization_note': 'Critical caches cleared instantly, comprehensive cache clearing in background'
            }
        }
        
        if errors:
            response_data['errors'] = errors
            response_data['message'] += f' ({len(errors)} notes/errors)'
        
        # Add cache performance data to response
        response_data['cache_cleared'] = True
        response_data['cache_performance'] = {
            'critical_keys_cleared': len(cache_keys_cleared),
            'critical_clear_time': f"{cache_clear_time:.3f}s",
            'comprehensive_clearing': 'in_background',
            'background_cache_keys': len(cache_keys_for_background),
            'types_cleared': cache_keys_cleared
        }
        
        # BACKGROUND AGGREGATION: Start monthly aggregation in background thread
        logger.info("🚀 BACKGROUND AGGREGATION: Starting monthly aggregation in background thread...")
        
        try:
            import threading
            from ..utils.utils import run_bulk_aggregation
            
            def background_aggregation():
                """Background thread function for monthly aggregation and cache clearing"""
                try:
                    logger.info(f"🧵 BACKGROUND THREAD: Starting aggregation for {attendance_date.year}-{attendance_date.month:02d}")
                    
                    # Run monthly aggregation
                    result = run_bulk_aggregation(tenant, attendance_date)
                    logger.info(f"🧵 BACKGROUND THREAD: Aggregation completed with status: {result['status']}")
                    
                    # Clear comprehensive caches in background
                    logger.info(f"🧵 BACKGROUND THREAD: Starting comprehensive cache clearing...")
                    cache_start_time = time.time()
                    
                    from django.core.cache import cache
                    cache_keys_cleared_count = 0
                    
                    for cache_key in cache_keys_for_background:
                        cache.delete(cache_key)
                        cache_keys_cleared_count += 1
                    
                    cache_time = time.time() - cache_start_time
                    logger.info(f"🧵 BACKGROUND THREAD: Cleared {cache_keys_cleared_count} cache keys in {cache_time:.3f}s")
                    
                except Exception as e:
                    logger.error(f"❌ BACKGROUND THREAD ERROR: {str(e)}")
            
            # Start background thread with daemon=True
            background_thread = threading.Thread(target=background_aggregation, daemon=True)
            background_thread.start()
            
            logger.info(f"🧵 BACKGROUND THREAD: Started with ID {background_thread.ident}")
            
            # Update response data
            response_data['automatic_aggregation'] = {
                'status': 'started_in_background',
                'method': 'thread_based',
                'thread_id': background_thread.ident,
                'message': 'Monthly aggregation started in background thread'
            }
            
        except Exception as thread_error:
            logger.error(f"❌ BACKGROUND THREAD FAILED TO START: {thread_error}")
            response_data['automatic_aggregation'] = {
                'status': 'failed_to_start',
                'method': 'thread_based',
                'error': str(thread_error)
            }
        
        return Response(response_data, status=200)
        
    except Exception as e:
        logger.error(f"Error in bulk update attendance: {str(e)}")
        return Response({"error": "Failed to update attendance"}, status=500)

# Clean replacement for the update_monthly_summaries_parallel function

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def update_monthly_summaries_parallel(request):
    """
    Asynchronous API for updating monthly summaries after bulk attendance upload.
    Returns immediately while processing summaries in background thread.
    
    Expected usage:
    1. Frontend calls this API after bulk attendance upload
    2. Returns success immediately 
    3. Processing happens in background thread using ULTRA-FAST bulk operations
    4. Cache is cleared immediately for instant UI updates
    """
    try:
        import threading
        from datetime import datetime
        from django.core.cache import cache
        
        # Get logger instance
        logger = logging.getLogger(__name__)
        
        start_time = time.time()
        
        tenant = getattr(request, 'tenant', None)
        if not tenant:
            return Response({"error": "No tenant found"}, status=400)
        
        date_str = request.data.get('date')
        employee_ids = request.data.get('employee_ids', [])
        
        if not date_str:
            return Response({"error": "Date is required"}, status=400)
        
        # Parse the date
        try:
            attendance_date = datetime.strptime(date_str, '%Y-%m-%d').date()
        except ValueError:
            return Response({"error": "Invalid date format. Use YYYY-MM-DD"}, status=400)
        
        # DEBUG: Initial request information
        logger.info(f"🚀 ASYNC SUMMARY: API called with parameters:")
        logger.info(f"   📅 Date: {date_str} (parsed as {attendance_date})")
        logger.info(f"   👥 Employee IDs count: {len(employee_ids)}")
        logger.info(f"   🏢 Tenant: {tenant.name} (ID: {tenant.id})")
        logger.info(f"   👥 Employee IDs (first 10): {employee_ids[:10]}{'...' if len(employee_ids) > 10 else ''}")
        
        print(f"🚀 ASYNC SUMMARY: API called - Date: {date_str}, Employee count: {len(employee_ids)}, Tenant: {tenant.name}")
        
        logger.info(f"🔄 ASYNC SUMMARY: Starting background monthly summary update for {len(employee_ids)} employees on {date_str}")
        
        # CLEAR ALL RELATED CACHES IMMEDIATELY for instant UI updates
        cache_start_time = time.time()
        cache_keys_to_clear = [
            f"payroll_overview_{tenant.id}",
            f"months_with_attendance_{tenant.id}",
            f"eligible_employees_{tenant.id}",
            f"eligible_employees_{tenant.id}_progressive",
            f"directory_data_{tenant.id}",
            f"attendance_all_records_{tenant.id}",
            f"attendance_log_{tenant.id}",
            f"attendance_tracker_{tenant.id}",
            f"monthly_attendance_summary_{tenant.id}_{attendance_date.year}_{attendance_date.month}",
            f"monthly_attendance_summary_{tenant.id}",
            f"dashboard_stats_{tenant.id}",
            f"employee_attendance_history_{tenant.id}",
            # CRITICAL: Add all_records cache variations that need immediate clearing
            f"attendance_all_records_{tenant.id}_this_month_None_None_None_None",
            f"attendance_all_records_{tenant.id}_last_6_months_None_None_None_None",
            f"attendance_all_records_{tenant.id}_last_12_months_None_None_None_None",
            f"attendance_all_records_{tenant.id}_last_5_years_None_None_None_None",
            f"frontend_charts_{tenant.id}",
        ]
        
        # Clear all cache keys
        for cache_key in cache_keys_to_clear:
            cache.delete(cache_key)
            
        # Clear any date-specific cache keys
        cache.delete(f"attendance_all_records_{tenant.id}_{date_str}")
        cache.delete(f"eligible_employees_{tenant.id}_{date_str}")
        
        # Clear custom date-based all_records cache keys that might exist
        cache.delete(f"attendance_all_records_{tenant.id}_custom_{attendance_date.month}_{attendance_date.year}_None_None")
        cache.delete(f"attendance_all_records_{tenant.id}_custom_range_None_None_{date_str}_None")
        
        cache_time = time.time() - cache_start_time
        logger.info(f"🗑️ ASYNC SUMMARY: Cleared {len(cache_keys_to_clear) + 4} cache keys in {cache_time:.3f}s")
        logger.info(f"🗑️ ASYNC SUMMARY: Cache keys cleared: {cache_keys_to_clear[:5]}{'...' if len(cache_keys_to_clear) > 5 else ''}")
        
        # Define background processing function using our optimized aggregation
        def process_summaries_background():
            """Use our optimized run_bulk_aggregation function for monthly summaries"""
            try:
                from ..utils.utils import run_bulk_aggregation
                logger.info(f"🧵 BACKGROUND THREAD: Starting monthly aggregation for {attendance_date.year}-{attendance_date.month:02d}")
                result = run_bulk_aggregation(tenant, attendance_date)
                logger.info(f"🧵 BACKGROUND THREAD: Monthly aggregation completed with status: {result['status']}")
            except Exception as e:
                logger.error(f"❌ BACKGROUND THREAD ERROR: {str(e)}")
        
        
        # Start background processing thread
        if employee_ids:
            logger.info(f"🧵 ASYNC SUMMARY: About to start background thread for {len(employee_ids)} employees")
            print(f"🧵 CONSOLE: About to start background thread for {len(employee_ids)} employees")  # Console fallback
            
            background_thread = threading.Thread(target=process_summaries_background, daemon=True)
            background_thread.start()
            
            logger.info(f"🧵 ASYNC SUMMARY: Background thread started successfully - Thread ID: {background_thread.ident}")
            logger.info(f"🧵 ASYNC SUMMARY: Thread is alive: {background_thread.is_alive()}")
            print(f"🧵 CONSOLE: Background thread started - ID: {background_thread.ident}, Alive: {background_thread.is_alive()}")
        else:
            logger.warning(f"⚠️ ASYNC SUMMARY: No employee IDs provided - skipping background processing")
            print(f"⚠️ CONSOLE: No employee IDs provided - skipping background processing")
        
        # Return immediately with success response
        total_time = time.time() - start_time
        
        response_data = {
            'message': f'✅ Monthly summary update started! Processing {len(employee_ids)} employees in background.',
            'status': 'success',
            'summary_update': {
                'employees_to_process': len(employee_ids),
                'date': date_str,
                'month': f"{attendance_date.year}-{attendance_date.month:02d}",
                'update_method': 'async_background',
                'processing_status': 'started'
            },
            'performance': {
                'response_time': f"{total_time:.3f}s",
                'cache_clear_time': f"{cache_time:.3f}s",
                'cache_keys_cleared': len(cache_keys_to_clear) + 2,
                'processing_mode': 'ultra_fast_background_thread'
            },
            'cache_cleared': True,
            'background_processing': True
        }
        
        logger.info(f"ASYNC SUMMARY: Returned response in {total_time:.3f}s, background processing started")
        
        return Response(response_data, status=200)
        
    except Exception as e:
        logger.error(f"Error in async monthly summary update: {str(e)}")
        return Response({"error": "Failed to start monthly summary update"}, status=500)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_eligible_employees_for_date(request):
    """
    PROGRESSIVE LOADING API - Load first 50 employees immediately, then lazy load the rest
    
    Two modes:
    1. initial=true: Returns first 50 employees instantly
    2. remaining=true: Returns all remaining employees for lazy loading
    
    PERFORMANCE IMPROVEMENTS:
    - Database-level slicing for instant first batch
    - Bulk attendance lookup with single query
    - Minimized Python loops and processing
    - Optimized data serialization
    
    EXPECTED: ~50ms for first 50, ~200ms for remaining
    """
    try:
        from datetime import datetime
        from django.db.models import Q
        from django.core.cache import cache
        
        # Performance timing
        start_time = time.time()
        
        tenant = getattr(request, 'tenant', None)
        if not tenant:
            return Response({"error": "No tenant found"}, status=400)
        
        date_str = request.query_params.get('date')
        if not date_str:
            return Response({"error": "Date parameter is required"}, status=400)
        
        # PROGRESSIVE LOADING PARAMETERS
        load_initial = request.query_params.get('initial', 'true').lower() == 'true'
        load_remaining = request.query_params.get('remaining', 'false').lower() == 'true'
        
        # Determine batch size and offset based on loading mode
        if load_initial and not load_remaining:
            # Mode 1: Load first 500 employees immediately (increased from 50)
            page_size = 500
            offset = 0
            cache_suffix = 'initial'
            load_mode = 'initial'
        elif load_remaining:
            # Mode 2: Load remaining employees (skip first 500)
            page_size = 2000  # Load all remaining at once
            offset = 500
            cache_suffix = 'remaining'
            load_mode = 'remaining'
        else:
            # Fallback: load first 500 (backward compatibility)
            page_size = 500
            offset = 0
            cache_suffix = 'initial'
            load_mode = 'fallback'
        
        try:
            target_date = datetime.strptime(date_str, '%Y-%m-%d').date()
        except ValueError:
            return Response({"error": "Invalid date format. Use YYYY-MM-DD"}, status=400)
        
        # Check cache first
        cache_key = f"eligible_employees_progressive_{tenant.id}_{date_str}_{cache_suffix}"
        use_cache = request.GET.get('no_cache', '').lower() != 'true'
        
        if use_cache:
            cached_data = cache.get(cache_key)
            if cached_data:
                cached_data['performance'] = {
                    'query_time': f"{(time.time() - start_time):.3f}s",
                    'cached': True,
                    'load_mode': 'initial' if load_initial and not load_remaining else 'remaining'
                }
                return Response(cached_data)
        
        # Get day of week for off-day checks
        day_of_week = target_date.weekday()
        day_names = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday']
        day_name = day_names[day_of_week]
        
        # Build off-day filter efficiently
        off_day_filters = {
            0: Q(off_monday=True),    # Monday
            1: Q(off_tuesday=True),   # Tuesday
            2: Q(off_wednesday=True), # Wednesday
            3: Q(off_thursday=True),  # Thursday
            4: Q(off_friday=True),    # Friday
            5: Q(off_saturday=True),  # Saturday
            6: Q(off_sunday=True),    # Sunday
        }
        off_day_filter = off_day_filters.get(day_of_week, Q())
        
        # PROGRESSIVE LOADING: Get total count once (cached for both requests)
        total_count_cache_key = f"total_eligible_count_{tenant.id}_{date_str}"
        total_count = cache.get(total_count_cache_key)
        
        if total_count is None:
            total_count = EmployeeProfile.objects.filter(
                tenant=tenant,
                is_active=True
            ).exclude(
                off_day_filter
            ).exclude(
                date_of_joining__gt=target_date
            ).count()
            # Cache total count for 5 minutes
            cache.set(total_count_cache_key, total_count, 300)
        
        # OPTIMIZATION 1: Database-level slicing for progressive loading
        eligible_employees_query = EmployeeProfile.objects.filter(
            tenant=tenant,
            is_active=True
        ).exclude(
            off_day_filter
        ).exclude(
            date_of_joining__gt=target_date
        ).only(
            # OPTIMIZATION 2: Only fetch required fields
            'employee_id', 'first_name', 'last_name', 'department',
            'shift_start_time', 'shift_end_time', 'is_active', 'date_of_joining'
        ).order_by('employee_id')[offset:offset + page_size]  # Progressive loading slice
        
        # OPTIMIZATION 4: Single bulk query for all attendance records  
        employee_ids = [emp.employee_id for emp in eligible_employees_query]
        
        # Bulk fetch attendance records in one query
        attendance_records = DailyAttendance.objects.filter(
            tenant=tenant,
            date=target_date,
            employee_id__in=employee_ids
        ).only(
            'employee_id', 'attendance_status', 'ot_hours', 'late_minutes', 'check_in', 'check_out'
        )
        
        # OPTIMIZATION 5: Build attendance lookup dictionary efficiently
        attendance_lookup = {
            record.employee_id: {
                'status': record.attendance_status,
                'ot_hours': float(record.ot_hours),
                'late_minutes': record.late_minutes,
                'check_in': record.check_in.strftime('%H:%M') if record.check_in else None,
                'check_out': record.check_out.strftime('%H:%M') if record.check_out else None,
            }
            for record in attendance_records
        }
        
        # OPTIMIZATION 4: Efficient data serialization with minimal processing
        eligible_employees = []
        for employee in eligible_employees_query:
            current_attendance = attendance_lookup.get(employee.employee_id, {})
            
            # Quick status determination
            if current_attendance:
                default_status = 'present' if current_attendance['status'] in ['PRESENT', 'PAID_LEAVE'] else 'absent'
            else:
                # No attendance record exists - leave unmarked
                default_status = None
            
            # Minimal data processing
            eligible_employees.append({
                'employee_id': employee.employee_id,
                'name': f"{employee.first_name} {employee.last_name}",
                'first_name': employee.first_name,
                'last_name': employee.last_name,
                'department': employee.department or 'General',
                'shift_start_time': employee.shift_start_time.strftime('%H:%M') if employee.shift_start_time else '09:00',
                'shift_end_time': employee.shift_end_time.strftime('%H:%M') if employee.shift_end_time else '18:00',
                'default_status': default_status,
                'current_attendance': current_attendance,
                'ot_hours': current_attendance.get('ot_hours', 0),
                'late_minutes': current_attendance.get('late_minutes', 0)
            })
        
        # PROGRESSIVE LOADING METADATA
        is_initial_load = load_initial and not load_remaining
        is_remaining_load = load_remaining
        remaining_count = max(0, total_count - 50) if is_initial_load else 0
        
        response_data = {
            'date': date_str,
            'day_name': day_name,
            'eligible_employees': eligible_employees,
            'progressive_loading': {
                'is_initial_load': is_initial_load,
                'is_remaining_load': is_remaining_load,
                'employees_in_batch': len(eligible_employees),
                'total_employees': total_count,
                'remaining_employees': remaining_count,
                'has_more': remaining_count > 0 if is_initial_load else False,
                'next_batch_url': f"/api/eligible-employees/?date={date_str}&remaining=true" if remaining_count > 0 and is_initial_load else None,
                'preserve_user_changes': True,  # Frontend should preserve user modifications
                'auto_trigger_remaining': is_initial_load and remaining_count > 0,  # Should auto-trigger background load
                'batch_offset': offset,
                'recommended_delay_ms': 100  # Suggested delay before background load
            },
            'total_count': len(eligible_employees),
            'performance': {
                'query_time': f"{(time.time() - start_time):.3f}s",
                'cached': False,
                'load_mode': 'initial' if is_initial_load else 'remaining',
                'batch_size': len(eligible_employees),
                'total_employees': total_count
            }
        }
        
        # Cache for 2 minutes
        if use_cache:
            cache.set(cache_key, response_data, 120)
        
        return Response(response_data)
        
    except Exception as e:
        logger.error(f"Error getting eligible employees: {str(e)}")
        return Response({"error": "Failed to get eligible employees"}, status=500)


class UploadAttendanceDataAPIView(APIView):
    """
    API endpoint for uploading monthly attendance summary data from Excel files
    Only accepts monthly summary format with columns: Employee ID, Name, Department, Present Days, Absent Days, OT Hours, Late Minutes
    """
    parser_classes = [MultiPartParser, FormParser]
    permission_classes = [IsAuthenticated]

    def post(self, request):
        import time
        start_time = time.time()
        
        try:
            import calendar
            import pandas as pd
            from datetime import datetime, date
            from django.db import transaction
            from ..models import Attendance, EmployeeProfile
            
            # Get tenant
            tenant = getattr(request, 'tenant', None)
            if not tenant:
                return Response({
                    'error': 'No tenant found for this request'
                }, status=status.HTTP_400_BAD_REQUEST)
            
            # Get file and parameters
            file_obj = request.FILES.get('file')
            month = request.data.get('month')
            year = request.data.get('year')
            
            if not file_obj:
                return Response({
                    'error': 'No file provided'
                }, status=status.HTTP_400_BAD_REQUEST)
            
            if not month or not year:
                return Response({
                    'error': 'Month and year are required'
                }, status=status.HTTP_400_BAD_REQUEST)
            
            # Validate file type
            if not (file_obj.name.endswith('.xlsx') or file_obj.name.endswith('.xls')):
                return Response({
                    'error': 'Unsupported file format. Please upload Excel (.xlsx, .xls) files only.'
                }, status=status.HTTP_400_BAD_REQUEST)
            
            try:
                # Read Excel file
                df = pd.read_excel(file_obj)
                
                # Check for monthly summary format (Employee ID and Working Days are optional)
                required_columns = ['Name', 'Department', 'Present Days', 'Absent Days', 'OT Hours', 'Late Minutes']
                optional_columns = ['Working Days']
                
                is_monthly_format = all(col in df.columns for col in required_columns)
                has_working_days = 'Working Days' in df.columns
                has_employee_id = False  # Employee ID no longer expected in template
                
                if not is_monthly_format:
                    return Response({
                        'error': f'Invalid file format. Expected monthly summary columns: {", ".join(required_columns)}',
                        'available_columns': list(df.columns),
                        'required_columns': required_columns,
                        'optional_columns': optional_columns
                    }, status=status.HTTP_400_BAD_REQUEST)
                
                # Process data
                records_created = 0
                records_updated = 0
                errors = []
                warnings = []
                
                # Process monthly summary format only
                # Auto-generate Employee IDs for entries that don't have them
                from ..utils.utils import generate_employee_id_bulk_optimized
                
                # OPTIMIZED: Use values() to reduce data transfer, iterator for memory efficiency
                existing_employees_by_name = {}
                for emp in EmployeeProfile.objects.filter(
                    tenant=tenant, is_active=True
                ).values('first_name', 'last_name', 'employee_id').iterator(chunk_size=1000):
                    key = f"{emp['first_name']} {emp['last_name']}".strip().lower()
                    existing_employees_by_name[key] = emp['employee_id']
                
                # Prepare data for Employee ID generation and matching
                employees_data = []
                final_employee_ids = []
                
                for index, row in df.iterrows():
                    name = str(row['Name']).strip()
                    department = str(row.get('Department', '')).strip()
                    
                    # Employee ID will be auto-generated based on name matching
                    provided_id = ''
                    
                    if provided_id:
                        # Use provided Employee ID
                        final_employee_ids.append(provided_id)
                    else:
                        # Try to match by name first
                        name_key = name.lower()
                        if name_key in existing_employees_by_name:
                            # Found existing employee by name
                            final_employee_ids.append(existing_employees_by_name[name_key])
                        else:
                            # Need to generate new Employee ID
                            employees_data.append({
                                'name': name,
                                'department': department
                            })
                            final_employee_ids.append(None)  # Will be filled after generation
                
                # Generate Employee IDs for new employees only
                if employees_data:
                    employee_id_mapping = generate_employee_id_bulk_optimized(employees_data, tenant.id)
                    
                    # Fill in the generated IDs
                    generated_index = 0
                    for i, emp_id in enumerate(final_employee_ids):
                        if emp_id is None:
                            final_employee_ids[i] = employee_id_mapping[generated_index]
                            generated_index += 1
                
                # Add final Employee IDs to the dataframe
                df['Final_Employee_ID'] = final_employee_ids
                
                # OPTIMIZED: Reuse existing_employees_by_name to avoid second query
                employee_ids = df['Final_Employee_ID'].dropna().unique()
                existing_employee_set = set(existing_employees_by_name.values())
                missing_employees = set(employee_ids) - existing_employee_set
                
                # If there are missing employees, collect their details and return for confirmation
                if missing_employees:
                    missing_employee_details = []
                    for index, row in df.iterrows():
                        employee_id = str(row['Final_Employee_ID']).strip()
                        if employee_id in missing_employees:
                            name = str(row['Name']).strip()
                            department = str(row.get('Department', '')).strip()
                            
                            # Split name into first and last name
                            name_parts = name.split(' ', 1)
                            first_name = name_parts[0] if name_parts else ''
                            last_name = name_parts[1] if len(name_parts) > 1 else ''
                            
                            # Clean up last name - avoid "nan" values
                            if not last_name or str(last_name).lower() in ['nan', 'none', '']:
                                last_name = ''
                            
                            missing_employee_details.append({
                                'employee_id': employee_id,
                                'name': name,
                                'first_name': first_name,
                                'last_name': last_name,
                                'department': department,
                                'row_number': index + 2  # Excel row number (accounting for header)
                            })
                    
                    return Response({
                        'error': 'Missing employees found',
                        'missing_employees': missing_employee_details,
                        'total_missing': len(missing_employee_details),
                        'message': f'Found {len(missing_employee_details)} employees that do not exist in the system. Please confirm to create them.'
                    }, status=status.HTTP_400_BAD_REQUEST)
                
                # ULTRA-OPTIMIZED: Use bulk operations instead of N queries
                from datetime import date
                attendance_date = date(int(year), int(month), 1)
                calendar_days = calendar.monthrange(int(year), int(month))[1]
                
                # Get existing attendance records for this period (single query)
                existing_attendance_dict = {
                    att['employee_id']: att['id']
                    for att in Attendance.objects.filter(
                        tenant=tenant,
                        date=attendance_date
                    ).values('employee_id', 'id')
                }
                
                # Prepare bulk create and update lists
                attendance_to_create = []
                attendance_to_update = []
                
                for index, row in df.iterrows():
                    try:
                        employee_id = str(row['Final_Employee_ID']).strip()
                        
                        # Process monthly summary format
                        name = str(row['Name']).strip()
                        department = str(row.get('Department', '')).strip()
                        
                        # Validate employee exists
                        if employee_id not in existing_employee_set:
                            errors.append(f'Row {index + 2}: Employee {employee_id} not found or inactive')
                            continue
                        
                        # Get numeric values
                        present_days = float(row.get('Present Days', 0)) if pd.notna(row.get('Present Days')) else 0
                        absent_days = float(row.get('Absent Days', 0)) if pd.notna(row.get('Absent Days')) else 0
                        ot_hours = float(row.get('OT Hours', 0)) if pd.notna(row.get('OT Hours')) else 0
                        late_minutes = int(row.get('Late Minutes', 0)) if pd.notna(row.get('Late Minutes')) else 0
                        
                        # OPTIMIZED: Use standard 30 days for bulk performance (DOJ calc in background if needed)
                        if has_working_days and pd.notna(row.get('Working Days')):
                            total_working_days = float(row.get('Working Days', 0))
                            if total_working_days <= 0:
                                total_working_days = 30
                        else:
                            total_working_days = 30
                        
                        attendance_data = {
                            'tenant': tenant,
                            'employee_id': employee_id,
                            'name': name,
                            'department': department,
                            'date': attendance_date,
                            'calendar_days': calendar_days,
                            'total_working_days': int(total_working_days),
                            'present_days': present_days,
                            'absent_days': absent_days,
                            'ot_hours': ot_hours,
                            'late_minutes': late_minutes
                        }
                        
                        if employee_id in existing_attendance_dict:
                            # Update existing record
                            attendance_data['id'] = existing_attendance_dict[employee_id]
                            attendance_to_update.append(Attendance(**attendance_data))
                            records_updated += 1
                        else:
                            # Create new record
                            attendance_to_create.append(Attendance(**attendance_data))
                            records_created += 1
                            
                    except Exception as e:
                        errors.append(f'Row {index + 2}: {str(e)}')
                
                # BULK OPERATIONS: 10-100x faster than individual saves!
                with transaction.atomic():
                    if attendance_to_create:
                        Attendance.objects.bulk_create(attendance_to_create, batch_size=100, ignore_conflicts=True)
                    
                    if attendance_to_update:
                        Attendance.objects.bulk_update(
                            attendance_to_update,
                            ['name', 'department', 'total_working_days', 'present_days', 
                             'absent_days', 'ot_hours', 'late_minutes', 'calendar_days'],
                            batch_size=100
                        )
                
                
                # Clear relevant caches
                from django.core.cache import cache
                cache_keys = [
                    f"payroll_overview_{tenant.id}",
                    f"attendance_all_records_{tenant.id}",
                    f"directory_data_{tenant.id}",
                    f"directory_data_full_{tenant.id}",  # Clear full directory cache
                    f"months_with_attendance_{tenant.id}"
                ]
                for key in cache_keys:
                    cache.delete(key)
                
                # Clear frontend charts cache (stats component)
                try:
                    cache.delete_pattern(f"frontend_charts_{tenant.id}_*")
                    logger.info(f"✨ Cleared frontend_charts pattern cache for tenant {tenant.id}")
                except AttributeError:
                    # Fallback: Clear common chart cache keys
                    chart_keys = [
                        f"frontend_charts_{tenant.id}_this_month_All_",
                        f"frontend_charts_{tenant.id}_last_6_months_All_",
                        f"frontend_charts_{tenant.id}_last_12_months_All_",
                        f"frontend_charts_{tenant.id}_last_5_years_All_"
                    ]
                    for key in chart_keys:
                        cache.delete(key)
                    logger.info(f"✨ Cleared frontend_charts fallback cache for tenant {tenant.id}")
                
                logger.info(f"✨ Cleared directory and charts cache for tenant {tenant.id} after attendance upload")
                
                # Calculate upload time
                upload_time = round((time.time() - start_time) * 1000, 2)  # milliseconds
                
                return Response({
                    'message': 'Attendance data uploaded successfully!',
                    'records_created': records_created,
                    'records_updated': records_updated,
                    'total_errors': len(errors),
                    'total_warnings': len(warnings),
                    'upload_time_ms': upload_time,
                    'errors': errors[:10],  # Show first 10 errors
                    'warnings': warnings[:10],  # Show first 10 warnings
                    'month': month,
                    'year': year,
                    'file_name': file_obj.name
                }, status=status.HTTP_201_CREATED)
                
            except Exception as e:
                return Response({
                    'error': f'Failed to process file: {str(e)}'
                }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
                
        except Exception as e:
            logger.error(f"Error in attendance upload: {str(e)}")
            return Response({
                'error': 'Upload failed. Please try again.'
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class UploadMonthlyAttendanceAPIView(APIView):
    """
    API endpoint for uploading monthly attendance summary data
    """
    permission_classes = [IsAuthenticated]

    def post(self, request):
        try:
            import calendar
            from datetime import date
            from django.db import transaction
            from ..models import Attendance
            
            # Get tenant
            tenant = getattr(request, 'tenant', None)
            if not tenant:
                return Response({
                    'error': 'No tenant found for this request'
                }, status=status.HTTP_400_BAD_REQUEST)
            
            # Get parameters
            month = request.data.get('month')
            year = request.data.get('year')
            data = request.data.get('data', [])
            
            if not month or not year:
                return Response({
                    'error': 'Month and year are required'
                }, status=status.HTTP_400_BAD_REQUEST)
            
            if not data:
                return Response({
                    'error': 'No attendance data provided'
                }, status=status.HTTP_400_BAD_REQUEST)
            
            created = 0
            failed = 0
            errors = []
            
            # Create attendance records
            attendance_records = []
            
            for index, record in enumerate(data):
                try:
                    employee_id = record.get('employee_id')
                    name = record.get('name')
                    
                    if not employee_id or not name:
                        errors.append(f'Record {index + 1}: Missing employee_id or name')
                        failed += 1
                        continue
                    
                    # Create attendance date (first day of the month)
                    attendance_date = date(int(year), int(month), 1)
                    
                    attendance_record = Attendance(
                        tenant=tenant,
                        employee_id=employee_id,
                        name=name,
                        department=record.get('department', ''),
                        date=attendance_date,
                        calendar_days=calendar.monthrange(int(year), int(month))[1],
                        total_working_days=record.get('total_working_days', 0),
                        present_days=record.get('present_days', 0),
                        absent_days=record.get('absent_days', 0),
                        ot_hours=record.get('ot_hours', 0),
                        late_minutes=record.get('late_minutes', 0)
                    )
                    
                    attendance_records.append(attendance_record)
                    created += 1
                    
                except Exception as e:
                    errors.append(f'Record {index + 1}: {str(e)}')
                    failed += 1
            
            # Bulk create records
            if attendance_records:
                with transaction.atomic():
                    Attendance.objects.bulk_create(attendance_records, ignore_conflicts=True)
            
            # Clear directory cache after successful upload
            from django.core.cache import cache
            cache_keys = [
                f"directory_data_{tenant.id}",
                f"directory_data_full_{tenant.id}",  # Clear full directory cache
                f"attendance_all_records_{tenant.id}"
            ]
            for key in cache_keys:
                cache.delete(key)
            
            # Clear frontend charts cache (stats component)
            try:
                cache.delete_pattern(f"frontend_charts_{tenant.id}_*")
                logger.info(f"✨ Cleared frontend_charts pattern cache for tenant {tenant.id}")
            except AttributeError:
                # Fallback: Clear common chart cache keys
                chart_keys = [
                    f"frontend_charts_{tenant.id}_this_month_All_",
                    f"frontend_charts_{tenant.id}_last_6_months_All_",
                    f"frontend_charts_{tenant.id}_last_12_months_All_",
                    f"frontend_charts_{tenant.id}_last_5_years_All_"
                ]
                for key in chart_keys:
                    cache.delete(key)
                logger.info(f"✨ Cleared frontend_charts fallback cache for tenant {tenant.id}")
            
            logger.info(f"✨ Cleared directory and charts cache for tenant {tenant.id} after monthly attendance upload")
            
            return Response({
                'message': 'Monthly attendance data uploaded successfully',
                'total_records': len(data),
                'created': created,
                'failed': failed,
                'errors': errors[:10],  # Show first 10 errors
                'month': month,
                'year': year
            }, status=status.HTTP_201_CREATED)
            
        except Exception as e:
            logger.error(f"Error in monthly attendance upload: {str(e)}")
            return Response({
                'error': 'Upload failed. Please try again.'
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class DownloadAttendanceTemplateAPIView(APIView):
    """
    API endpoint for downloading attendance template
    """
    permission_classes = [IsAuthenticated]

    def get(self, request):
        try:
            from openpyxl import Workbook
            from django.http import HttpResponse
            
            # Create workbook and worksheet
            wb = Workbook()
            ws = wb.active
            ws.title = "Attendance Template"
            
            # Define headers for monthly attendance format (Employee ID is optional)
            headers = [
                'Name', 'Department', 'Present Days', 'Absent Days', 'Working Days', 'OT Hours', 'Late Minutes'
            ]
            
            # Add headers to worksheet
            for col, header in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=header)
            
            # Add sample data for monthly format (Employee IDs will be auto-generated based on name matching)
            sample_data = [
                ['John Doe', 'Engineering', 22, 3, 25, 15.5, 45],
                ['Jane Smith', 'HR', 20, 5, 25, 8.0, 30],
                ['Bob Johnson', 'Sales', 18, 7, 25, 12.0, 60],
                ['Alice Brown', 'Marketing', 21, 4, 25, 5.5, 25],
                ['Charlie Wilson', 'Finance', 19, 6, 25, 10.0, 40]
            ]
            
            for row_idx, row_data in enumerate(sample_data, 2):
                for col_idx, value in enumerate(row_data, 1):
                    ws.cell(row=row_idx, column=col_idx, value=value)
            
            # Add instructions
            ws.cell(row=8, column=1, value="Instructions:")
            ws.cell(row=9, column=1, value="1. Name: Full name of the employee (Employee ID will be auto-generated)")
            ws.cell(row=10, column=1, value="2. Department: Employee's department")
            ws.cell(row=11, column=1, value="3. Present Days: Number of days employee was present")
            ws.cell(row=12, column=1, value="4. Absent Days: Number of days employee was absent")
            ws.cell(row=13, column=1, value="5. Working Days: Total working days (optional - will calculate from DOJ if not provided)")
            ws.cell(row=14, column=1, value="6. OT Hours: Total overtime hours (decimal, e.g., 15.5)")
            ws.cell(row=15, column=1, value="7. Late Minutes: Total late minutes (integer, e.g., 45)")
            ws.cell(row=16, column=1, value="8. This is a MONTHLY SUMMARY format - no individual dates needed")
            
            # Create response
            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename=attendance_template.xlsx'
            wb.save(response)
            
            return response
            
        except Exception as e:
            return Response({
                'error': f'Failed to generate template: {str(e)}'
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
