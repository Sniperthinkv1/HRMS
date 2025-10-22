# multi_tenant.py
# Contains views related to multi-tenant operations:
# - TenantViewSet
# - UploadSalaryDataAPIView
# - DownloadTemplateAPIView

from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.parsers import MultiPartParser, FormParser
from rest_framework import status, viewsets
from rest_framework.permissions import IsAuthenticated
from django.db import transaction
from django.http import HttpResponse
import pandas as pd
from datetime import datetime
import logging
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

from ..models import (
    Tenant,
    SalaryData,
    EmployeeProfile,
)
from ..serializers import (
    TenantSerializer,
)
from ..utils.permissions import IsSuperUser
from ..utils.utils import (
    clean_decimal_value,
    clean_int_value,
    is_valid_name,
    validate_excel_columns,
    generate_employee_id,
)

TEMPLATE_COLUMNS = [
    "NAME",
    "SALARY",
    "ABSENT",
    "Present Days",
    "Working Days",
    "SL W/O OT",
    "OT",
    "HOUR RS",
    "OT CHARGES",  # Renamed from CHARGES for clarity
    "LATE",
    "LATE CHARGE",  # Renamed from CHARGE for clarity
    "AMT",
    "SAL+OT",
    "25TH ADV",
    "OLD ADV",
    "NETT PAYABLE",
    "Department",
    "Total old ADV",
    "Balnce Adv",
    "INCENTIVE",
    "TDS",
    "SAL-TDS",
    "ADVANCE",
]

# Initialize logger
logger = logging.getLogger(__name__)

class TenantViewSet(viewsets.ModelViewSet):
    """
    Tenant management for super admins only
    """

    queryset = Tenant.objects.all()
    serializer_class = TenantSerializer
    permission_classes = [IsSuperUser]


class UploadSalaryDataAPIView(APIView):
    """
    Optimized salary data upload using exact template structure with bulk operations
    Requires existing employees only - no auto-creation (like attendance upload)
    """

    parser_classes = [MultiPartParser, FormParser]
    permission_classes = [IsAuthenticated]

    def post(self, request, *args, **kwargs):
        try:
            # Get tenant from request
            tenant = getattr(request, "tenant", None)
            if not tenant:

                return Response(
                    {
                        "error": "No tenant found. Please ensure you're accessing the correct subdomain."
                    },
                    status=status.HTTP_400_BAD_REQUEST,
                )

            excel_file = request.FILES.get("file")

            if not excel_file:

                return Response(
                    {"error": "No file uploaded"}, status=status.HTTP_400_BAD_REQUEST
                )

            # Validate file extension

            if not excel_file.name.endswith((".xlsx", ".xls")):

                return Response(
                    {"error": "Please upload an Excel file (.xlsx or .xls)"},
                    status=status.HTTP_400_BAD_REQUEST,
                )

            try:

                # Read Excel file with NaN handling

                df = pd.read_excel(excel_file)

                # Map alternative column names to standard template names
                column_mapping = {
                    'Basic Salary': 'SALARY',
                    'basic_salary': 'SALARY',
                    'BasicSalary': 'SALARY',
                    'basic salary': 'SALARY',
                }
                
                # Rename columns based on mapping (case-insensitive)
                df_columns = df.columns.tolist()
                rename_dict = {}
                for col in df_columns:
                    if col in column_mapping:
                        rename_dict[col] = column_mapping[col]
                
                if rename_dict:
                    df = df.rename(columns=rename_dict)

                # Replace all NaN values with appropriate defaults before processing

                df = df.fillna(
                    {
                        "NAME": "",
                        "SALARY": 0,
                        "ABSENT": 0,
                        "Present Days": 0,
                        "SL W/O OT": 0,
                        "OT": 0,
                        "HOUR RS": 0,
                        "OT CHARGES": 0,  # Renamed from CHARGES
                        "LATE": 0,
                        "LATE CHARGE": 0,  # Renamed from CHARGE
                        "AMT": 0,
                        "SAL+OT": 0,
                        "25TH ADV": 0,
                        "OLD ADV": 0,
                        "NETT PAYABLE": 0,
                        "Department": "",
                        "Total old ADV": 0,
                        "Balnce Adv": 0,
                        "INCENTIVE": 0,
                        "TDS": 0,
                        "SAL-TDS": 0,
                        "ADVANCE": 0,
                        "Working Days": 0,  # Optional column
                    }
                )

                # Validate columns (Employee ID and Working Days are optional)
                required_columns = [col for col in TEMPLATE_COLUMNS if col not in ["Working Days"]]
                optional_columns = ["Working Days"]
                
                is_valid, error_message = validate_excel_columns(
                    df.columns.tolist(), required_columns, optional_columns
                )

                if not is_valid:

                    return Response(
                        {"error": error_message}, status=status.HTTP_400_BAD_REQUEST
                    )

                # Get month/year from request

                selected_month = request.data.get("month")

                selected_year = request.data.get("year")

                if not selected_month or not selected_year:

                    return Response(
                        {"error": "Please select month and year for the uploaded data"},
                        status=status.HTTP_400_BAD_REQUEST,
                    )

                # Filter out invalid rows BEFORE processing

                valid_rows = df[df["NAME"].apply(is_valid_name)]

                if len(valid_rows) == 0:

                    return Response(
                        {"error": "No valid employee names found in the uploaded file"},
                        status=status.HTTP_400_BAD_REQUEST,
                    )

                # Process data with bulk operations for maximum performance

                records_created = 0

                records_updated = 0

                errors = []

                # Collect all data for bulk operations

                salary_records_to_create = []

                salary_records_to_update = []

                employee_profiles_to_create = []

                # Get existing salary records for this period to avoid duplicates

                existing_salary_dict = {}

                existing_salaries = SalaryData.objects.filter(
                    tenant=tenant, year=int(selected_year), month=selected_month
                ).values("employee_id", "id")

                for salary in existing_salaries:

                    existing_salary_dict[salary["employee_id"]] = salary["id"]

                # Auto-generate Employee IDs for entries that don't have them
                from ..utils.utils import generate_employee_id_bulk_optimized
                
                # OPTIMIZED: Use values() to reduce data transfer, single query
                existing_employees_by_name = {}
                for emp in EmployeeProfile.objects.filter(
                    tenant=tenant, is_active=True
                ).values('first_name', 'last_name', 'employee_id').iterator(chunk_size=1000):
                    key = f"{emp['first_name']} {emp['last_name']}".strip().lower()
                    existing_employees_by_name[key] = emp['employee_id']
                
                # Prepare data for Employee ID generation and matching
                employees_data = []
                final_employee_ids = []
                
                for index, row in valid_rows.iterrows():
                    name = str(row['NAME']).strip()
                    department = str(row.get('Department', '')).strip()
                    
                    # Check if we have a provided Employee ID (no longer expected in template)
                    provided_id = ''
                    
                    if provided_id:
                        # Use provided Employee ID
                        final_employee_ids.append(provided_id)
                    else:
                        # Try to match by name first
                        name_key = name.lower()
                        if name_key in existing_employees_by_name:
                            # Found existing employee by name
                            final_employee_ids.append(existing_employees_by_name[name_key])
                        else:
                            # Need to generate new Employee ID
                            employees_data.append({
                                'name': name,
                                'department': department
                            })
                            final_employee_ids.append(None)  # Will be filled after generation
                
                # Generate Employee IDs for new employees only
                if employees_data:
                    employee_id_mapping = generate_employee_id_bulk_optimized(employees_data, tenant.id)
                    
                    # Fill in the generated IDs
                    generated_index = 0
                    for i, emp_id in enumerate(final_employee_ids):
                        if emp_id is None:
                            final_employee_ids[i] = employee_id_mapping[generated_index]
                            generated_index += 1
                
                # Add final Employee IDs to the dataframe
                valid_rows = valid_rows.copy()
                valid_rows['Final_Employee_ID'] = final_employee_ids
                
                # OPTIMIZED: Reuse existing_employees_by_name instead of second query
                existing_employee_set = set(existing_employees_by_name.values())

                # Check for missing employees first
                employee_ids_in_file = valid_rows['Final_Employee_ID'].dropna().unique()
                missing_employees = set(employee_ids_in_file) - existing_employee_set
                
                # If there are missing employees, collect their details and return for confirmation
                if missing_employees:
                    missing_employee_details = []
                    for index, row in valid_rows.iterrows():
                        employee_id = str(row['Final_Employee_ID']).strip()
                        if employee_id in missing_employees:
                            name = str(row.get("NAME", "")).strip()
                            department = str(row.get("Department", "")).strip()
                            
                            # Split name into first and last name
                            name_parts = name.split(' ', 1)
                            first_name = name_parts[0] if name_parts else ''
                            last_name = name_parts[1] if len(name_parts) > 1 else ''
                            
                            # Clean up last name - avoid "nan" values
                            if not last_name or str(last_name).lower() in ['nan', 'none', '']:
                                last_name = ''
                            
                            # Get basic salary from Excel - crucial for employee profile
                            basic_salary = clean_decimal_value(row.get("SALARY", 0))
                            
                            missing_employee_details.append({
                                'employee_id': employee_id,
                                'name': name,
                                'first_name': first_name,
                                'last_name': last_name,
                                'department': department,
                                'basic_salary': float(basic_salary) if basic_salary else 0,  # Pass salary from Excel
                                'row_number': index + 2  # Excel row number (accounting for header)
                            })
                    
                    return Response({
                        'error': 'Missing employees found',
                        'missing_employees': missing_employee_details,
                        'total_missing': len(missing_employee_details),
                        'message': f'Found {len(missing_employee_details)} employees that do not exist in the system. Please confirm to create them.'
                    }, status=status.HTTP_400_BAD_REQUEST)

                # Prepare bulk data

                for index, row in valid_rows.iterrows():

                    try:
                        # Get employee ID from Excel (use final employee ID)
                        employee_id = str(row['Final_Employee_ID']).strip()

                        # Prepare salary data
                        # Use Working Days if provided, otherwise calculate from employee profile
                        working_days = clean_int_value(row.get("Working Days", 0))
                        if working_days > 0:
                            days_value = working_days
                        else:
                            # FIXED: Always use 30 working days
                            days_value = 30
                            
                            # COMMENTED OUT: Complex DOJ-based working days calculation
                            # # Calculate working days from employee profile (DOJ and off days)
                            # try:
                            #     from ..services.salary_service import SalaryCalculationService
                            #     employee = EmployeeProfile.objects.get(tenant=tenant, employee_id=employee_id, is_active=True)
                            #     month_names = ['JAN', 'FEB', 'MAR', 'APR', 'MAY', 'JUN', 'JUL', 'AUG', 'SEP', 'OCT', 'NOV', 'DEC']
                            #     days_value = SalaryCalculationService._calculate_employee_working_days(employee, int(selected_year), month_names[int(selected_month) - 1])
                            # except EmployeeProfile.DoesNotExist:
                            #     # Fallback: use calendar working days
                            #     import calendar
                            #     from datetime import date
                            #     days_in_month = calendar.monthrange(int(selected_year), int(selected_month))[1]
                            #     days_value = 0
                            #     for day in range(1, days_in_month + 1):
                            #         date_obj = date(int(selected_year), int(selected_month), day)
                            #         if date_obj.weekday() < 5:  # Monday-Friday
                            #             days_value += 1
                            # except Exception as e:
                            #     # Fallback: use calendar working days
                            #     import calendar
                            #     from datetime import date
                            #     days_in_month = calendar.monthrange(int(selected_year), int(selected_month))[1]
                            #     days_value = 0
                            #     for day in range(1, days_in_month + 1):
                            #         date_obj = date(int(selected_year), int(selected_month), day)
                            #         if date_obj.weekday() < 5:  # Monday-Friday
                            #             days_value += 1

                        salary_data = {
                            "tenant": tenant,
                            "employee_id": employee_id,
                            "year": int(selected_year),
                            "month": selected_month,
                            "name": str(row["NAME"]).strip(),
                            "salary": clean_decimal_value(row["SALARY"]),
                            "absent": clean_int_value(row["ABSENT"]),
                            "days": days_value,
                            "sl_wo_ot": clean_decimal_value(row["SL W/O OT"]),
                            "ot": clean_decimal_value(row["OT"]),
                            "hour_rs": clean_decimal_value(row["HOUR RS"]),
                            "charges": clean_decimal_value(row["OT CHARGES"]),  # Updated column name
                            "late": clean_int_value(row["LATE"]),
                            "charge": clean_decimal_value(row["LATE CHARGE"]),  # Updated column name
                            "amt": clean_decimal_value(row["AMT"]),
                            "sal_ot": clean_decimal_value(row["SAL+OT"]),
                            "adv_25th": clean_decimal_value(row["25TH ADV"]),
                            "old_adv": clean_decimal_value(row["OLD ADV"]),
                            "nett_payable": clean_decimal_value(row["NETT PAYABLE"]),
                            "department": str(row.get("Department", "")).strip(),
                            "total_old_adv": clean_decimal_value(row["Total old ADV"]),
                            "balnce_adv": clean_decimal_value(row["Balnce Adv"]),
                            "incentive": clean_decimal_value(row["INCENTIVE"]),
                            "tds": clean_decimal_value(row["TDS"]),
                            "sal_tds": clean_decimal_value(row["SAL-TDS"]),
                            "advance": clean_decimal_value(row["ADVANCE"]),
                            "date": datetime(
                                int(selected_year),
                                self._get_month_number(selected_month),
                                1,
                            ).date(),
                        }

                        if employee_id in existing_salary_dict:

                            # Update existing record

                            salary_data["id"] = existing_salary_dict[employee_id]

                            salary_records_to_update.append(SalaryData(**salary_data))

                            records_updated += 1

                        else:

                            # Create new record

                            salary_records_to_create.append(SalaryData(**salary_data))

                            records_created += 1

                        # No employee profile creation - employee must exist (like attendance upload)

                    except Exception as e:

                        errors.append(f"Row {index + 2}: {str(e)}")

                # Perform bulk operations

                with transaction.atomic():

                    # Deduplicate salary records to create (in case same employee appears multiple times in upload)
                    if salary_records_to_create:
                        # Create a dictionary to deduplicate by (employee_id, year, month)
                        unique_salary_records = {}
                        for record in salary_records_to_create:
                            key = (record.employee_id, record.year, record.month)
                            if key not in unique_salary_records:
                                unique_salary_records[key] = record
                            else:
                                # If duplicate found, keep the last one (most recent data)
                                unique_salary_records[key] = record
                        
                        # Convert back to list
                        salary_records_to_create = list(unique_salary_records.values())
                        records_created = len(salary_records_to_create)

                    # Bulk create new salary records

                    if salary_records_to_create:

                        SalaryData.objects.bulk_create(
                            salary_records_to_create, batch_size=100
                        )

                    # Bulk update existing salary records

                    if salary_records_to_update:

                        SalaryData.objects.bulk_update(
                            salary_records_to_update,
                            [
                                "name",
                                "salary",
                                "absent",
                                "days",
                                "sl_wo_ot",
                                "ot",
                                "hour_rs",
                                "charges",
                                "late",
                                "charge",
                                "amt",
                                "sal_ot",
                                "adv_25th",
                                "old_adv",
                                "nett_payable",
                                "department",
                                "total_old_adv",
                                "balnce_adv",
                                "incentive",
                                "tds",
                                "sal_tds",
                                "advance",
                                "date",
                            ],
                            batch_size=100,
                        )
                    # Create or update PayrollPeriod for the uploaded data
                    from ..services.salary_service import SalaryCalculationService
                    from ..models import PayrollPeriod, DataSource
                    from django.core.cache import cache
                    from decimal import Decimal
                    import calendar
                    
                    # Create PayrollPeriod with UPLOADED data source
                    payroll_period, period_created = PayrollPeriod.objects.get_or_create(
                        tenant=tenant,
                        year=int(selected_year),
                        month=selected_month,
                        defaults={
                            'data_source': DataSource.UPLOADED,
                            'working_days_in_month': len([d for d in range(1, calendar.monthrange(int(selected_year), self._get_month_number(selected_month))[1] + 1)
                                                          if calendar.weekday(int(selected_year), self._get_month_number(selected_month), d) < 5]),
                            'tds_rate': Decimal('5.00')
                        }
                    )
                    
                    # If period already exists, update data source to UPLOADED
                    if not period_created:
                        payroll_period.data_source = DataSource.UPLOADED
                        payroll_period.save()
                    
                    # Clear payroll overview cache to show new data immediately
                    cache_key = f"payroll_overview_{tenant.id}"
                    cache.delete(cache_key)
                    
                    # Clear frontend charts cache to refresh dashboard immediately
                    try:
                        # Try to clear all frontend charts cache variations
                        cache.delete_pattern(f"frontend_charts_{tenant.id}_*")
                    except AttributeError:
                        # Fallback: Clear specific common cache keys
                        cache.delete(f"frontend_charts_{tenant.id}")
                    
                    # Also clear any related caches
                    cache.delete(f"payroll_period_detail_{payroll_period.id}")
                    cache.delete(f"payroll_summary_{payroll_period.id}")
                    cache.delete(f"directory_data_{tenant.id}")
                    cache.delete(f"directory_data_full_{tenant.id}")  # Clear full directory cache
                    
                    logger.info(f"✨ Cleared payroll, charts, and directory cache for tenant {tenant.id} after salary upload")
                    
                    # ✨ BACKGROUND SYNC: Aggregate chart data in background thread
                    # This avoids blocking the response while processing chart aggregation
                    from ..utils.chart_sync import sync_chart_data_batch_async
                    sync_chart_data_batch_async(tenant, int(selected_year), selected_month, source='excel')
                    logger.info(f"📊 Triggered background chart aggregation for {selected_month} {selected_year}")
                    
                    # ✨ AUTOMATIC PAYROLL CALCULATION (BACKGROUND THREAD):
                    # Process uploaded salary data into CalculatedSalary without blocking the request
                    import threading
                    import time
                    
                    def _run_payroll_calculation_async(tenant_id: int, year: int, month: str):
                        from django.db import connection
                        from django.db import transaction
                        try:
                            # Small delay to ensure upload transaction is committed
                            time.sleep(1)
                            
                            from ..models import Tenant, SalaryData, CalculatedSalary, DataSource
                            from decimal import Decimal
                            from datetime import date
                            
                            _tenant = Tenant.objects.get(id=tenant_id)
                            logger.info(f"💰 [BG] Starting payroll calculation for {month} {year}")
                            
                            # Get uploaded salary data
                            salary_data = SalaryData.objects.filter(
                                tenant_id=tenant_id, year=year, month=month
                            )
                            
                            if not salary_data.exists():
                                logger.warning(f"💰 [BG] No SalaryData found for {month} {year}")
                                return
                            
                            # Get or create payroll period
                            from ..models import PayrollPeriod
                            period = PayrollPeriod.objects.get(
                                tenant_id=tenant_id, year=year, month=month
                            )
                            
                            with transaction.atomic():
                                for sd in salary_data:
                                    # Create CalculatedSalary record with Excel values
                                    calculated_salary = CalculatedSalary(
                                        tenant_id=tenant_id,
                                        payroll_period=period,
                                        employee_id=sd.employee_id,
                                        employee_name=sd.name,
                                        department=sd.department or 'General',
                                        basic_salary=sd.salary or Decimal('0'),
                                        basic_salary_per_hour=sd.hour_rs or Decimal('0'),
                                        basic_salary_per_minute=sd.charge or Decimal('0'),
                                        employee_ot_rate=sd.hour_rs or Decimal('0'),
                                        employee_tds_rate=sd.tds or Decimal('0'),
                                        total_working_days=int((sd.days or 0) + (sd.absent or 0)),
                                        present_days=Decimal(str(sd.days or 0)),
                                        absent_days=Decimal(str(sd.absent or 0)),
                                        ot_hours=sd.ot or Decimal('0'),
                                        late_minutes=int(sd.late or 0),
                                        salary_for_present_days=sd.sl_wo_ot or Decimal('0'),
                                        ot_charges=sd.charges or Decimal('0'),
                                        late_deduction=sd.amt or Decimal('0'),
                                        incentive=sd.incentive or Decimal('0'),
                                        gross_salary=sd.sal_ot or Decimal('0'),
                                        tds_amount=sd.tds or Decimal('0'),
                                        salary_after_tds=sd.sal_tds or Decimal('0'),
                                        total_advance_balance=sd.total_old_adv or Decimal('0'),
                                        advance_deduction_amount=sd.advance or Decimal('0'),
                                        advance_deduction_editable=True,
                                        remaining_advance_balance=sd.balnce_adv or Decimal('0'),
                                        net_payable=sd.nett_payable or Decimal('0'),
                                        data_source=DataSource.UPLOADED,
                                        is_paid=True,
                                        payment_date=date.today(),
                                    )
                                    
                                    # Skip auto-calculation
                                    calculated_salary._skip_auto_calc = True
                                    calculated_salary.save()
                                    
                            logger.info(f"💰 [BG] Created {salary_data.count()} CalculatedSalary records marked as paid")

                            # Clear caches so frontend reflects paid status immediately
                            try:
                                from excel_data.services.cache_service import invalidate_payroll_payment_caches
                                
                                cache_result = invalidate_payroll_payment_caches(
                                    tenant=tenant, 
                                    period_id=period.id,
                                    reason="uploaded_salary_data_marked_paid"
                                )
                                
                                if cache_result['success']:
                                    logger.info(f"🧹 [BG] Cache invalidation successful: {cache_result['cleared_count']} keys cleared for tenant {tenant_id}, period {period.id}")
                                else:
                                    logger.warning(f"⚠️ [BG] Cache invalidation failed: {cache_result.get('error', 'Unknown error')}")
                            except Exception as cache_exc:
                                logger.warning(f"⚠️ [BG] Cache clear failed: {cache_exc}")
                            
                        except Exception as exc:
                            logger.error(f"❌ [BG] Automatic payroll calculation failed: {exc}")
                        finally:
                            # Ensure DB connection is closed in thread
                            connection.close()

                    threading.Thread(
                        target=_run_payroll_calculation_async,
                        args=(tenant.id, int(selected_year), selected_month),
                        daemon=True
                    ).start()

                return Response(
                    {
                        "message": "Upload completed successfully",
                        "records_created": records_created,
                        "records_updated": records_updated,
                        "total_processed": len(valid_rows),
                        "total_rows_in_file": len(df),
                        "errors": errors[:10],  # Show first 10 errors only
                        "total_errors": len(errors),
                        "payroll_period_id": payroll_period.id,
                        "period_created": period_created,
                        "data_source": payroll_period.data_source,
                        "cache_cleared": True,
                        "automatic_payroll_processing": True,
                        "payroll_automatically_marked_as_paid": True
                    },
                    status=status.HTTP_200_OK,
                )

            except Exception as e:

                return Response(
                    {"error": f"Error processing Excel file: {str(e)}"},
                    status=status.HTTP_400_BAD_REQUEST,
                )

        except Exception as e:

            return Response(
                {"error": f"Upload failed: {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )

    def _get_month_number(self, month_name):
        """Convert month name to number"""

        months = {
            "JANUARY": 1,
            "JAN": 1,
            "FEBRUARY": 2,
            "FEB": 2,
            "MARCH": 3,
            "MAR": 3,
            "APRIL": 4,
            "APR": 4,
            "MAY": 5,
            "JUNE": 6,
            "JUN": 6,
            "JULY": 7,
            "JUL": 7,
            "AUGUST": 8,
            "AUG": 8,
            "SEPTEMBER": 9,
            "SEP": 9,
            "OCTOBER": 10,
            "OCT": 10,
            "NOVEMBER": 11,
            "NOV": 11,
            "DECEMBER": 12,
            "DEC": 12,
        }

        return months.get(month_name.upper(), 1)


class DownloadTemplateAPIView(APIView):
    """

    Download the Excel template for salary data upload

    """

    permission_classes = [IsAuthenticated]

    def get(self, request):

        # Create a new workbook and worksheet
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Salary Data Template"
        # Define styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(
            start_color="366092", end_color="366092", fill_type="solid"
        )
        border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
        center_alignment = Alignment(horizontal="center", vertical="center")

        # Add headers

        for col_idx, column in enumerate(TEMPLATE_COLUMNS, 1):
            cell = ws.cell(row=1, column=col_idx, value=column)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = center_alignment

        # Add sample data

        sample_data = [
            [
                "John Doe",  # NAME (Employee ID will be auto-generated)
                50000,       # SALARY
                2,           # ABSENT
                28,          # Present Days
                25,          # Working Days
                45000,       # SL W/O OT
                5,           # OT
                208.33,      # HOUR RS
                1041.65,     # OT CHARGES (Overtime payment)
                30,          # LATE
                6.94,        # LATE CHARGE (Late deduction)
                208.33,      # AMT
                46041.65,    # SAL+OT
                5000,        # 25TH ADV
                2000,        # OLD ADV
                39041.65,    # NETT PAYABLE
                "IT",        # Department
                10000,       # Total old ADV
                8000,        # Balnce Adv
                1000,        # INCENTIVE
                500,         # TDS
                38541.65,    # SAL-TDS
                0,           # ADVANCE
            ],
            [
                "Jane Smith", # NAME (Employee ID will be auto-generated)
                45000,        # SALARY
                1,            # ABSENT
                29,           # Present Days
                25,           # Working Days
                43500,        # SL W/O OT
                3,            # OT
                187.5,        # HOUR RS
                562.5,        # OT CHARGES (Overtime payment)
                15,           # LATE
                6.25,         # LATE CHARGE (Late deduction)
                93.75,        # AMT
                44062.5,      # SAL+OT
                4000,         # 25TH ADV
                1500,         # OLD ADV
                38562.5,      # NETT PAYABLE
                "HR",         # Department
                8000,         # Total old ADV
                6500,         # Balnce Adv
                800,          # INCENTIVE
                400,          # TDS
                38162.5,      # SAL-TDS
                0,            # ADVANCE
            ],
        ]

        for row_idx, row_data in enumerate(sample_data, 2):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = border
                if col_idx > 1:  # Numeric columns
                    cell.alignment = Alignment(horizontal="right")

        # Adjust column widths

        for col_idx, column in enumerate(TEMPLATE_COLUMNS, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = max(
                len(column) + 2, 12
            )

        # Create response

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = (
            'attachment; filename="salary_data_template.xlsx"'
        )
        wb.save(response)
        return response
